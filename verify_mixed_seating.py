import openpyxl

wb = openpyxl.load_workbook('Agent Schedules (Dec 1st - 21st)_20251218_151404_seating_arrangement.xlsx')

print("Sample from each sheet:")
print("=" * 80)

for sheet_name in ['Morning Seating', 'Night Seating', 'Mixed Seating', 'Full Seating']:
    ws = wb[sheet_name]
    print(f"\n{sheet_name} - First 3 data rows:")
    for i in range(2, 5):
        area = ws[f'A{i}'].value
        station = ws[f'B{i}'].value
        date_col_c = ws[f'C{i}'].value
        print(f"  Station {station:>2} ({area}): {date_col_c}")

# Check for "name / name" duplicates in Full Seating
print("\n" + "=" * 80)
print("Checking Full Seating for duplicate agent entries (e.g., 'John / John'):")
ws = wb['Full Seating']
duplicate_count = 0
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3):
    for cell in row:
        if cell.value and ' / ' in str(cell.value):
            parts = str(cell.value).split(' / ')
            if len(parts) == 2 and parts[0].strip() == parts[1].strip():
                duplicate_count += 1
                print(f"  Found: {cell.value}")
                if duplicate_count >= 5:
                    break
    if duplicate_count >= 5:
        break

if duplicate_count == 0:
    print("  ✓ No duplicate agent entries found!")
else:
    print(f"  Found {duplicate_count} duplicate entries")

# Show some combined entries
print("\n" + "=" * 80)
print("Sample combined entries in Full Seating:")
count = 0
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3):
    for cell in row:
        if cell.value and ' / ' in str(cell.value):
            print(f"  {cell.value}")
            count += 1
            if count >= 5:
                break
    if count >= 5:
        break
