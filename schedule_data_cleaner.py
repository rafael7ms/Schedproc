import pandas as pd
import os
import sys
import logging
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from difflib import get_close_matches

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler('schedule_processor.log')
    ]
)
logger = logging.getLogger(__name__)

class ScheduleProcessor:
    def __init__(self, auto_correct: bool = False):
        self.auto_correct = auto_correct
        self.regular_shifts = [
            ('05:00', '14:00'),
            ('06:00', '15:00'),
            ('07:00', '16:00'),
            ('10:00', '19:00'),
            ('11:00', '20:00'),
            ('14:00', '22:00'),
            ('15:00', '23:00'),
            ('15:30', '23:30'),
            ('16:00', '00:00'),
            ('17:00', '01:00')
        ]

    def clean_id(self, id_value):
        """Clean ID by removing whitespace and standardizing format"""
        if pd.isna(id_value):
            return None
        return str(id_value).strip()

    def find_best_match(self, target_id, id_list, threshold=0.9):
        """Find best matching ID using fuzzy matching"""
        if not id_list:
            return None
        matches = get_close_matches(target_id, id_list, n=1, cutoff=threshold)
        return matches[0] if matches else None

    def extract_column_name(self, name: str) -> str:
        """Extract the actual column name from prefixed names"""
        name = str(name).strip().lower()
        column_mapping = {
            'id': 'ID',
            'employeeid': 'ID',
            'agentid': 'ID',
            'firstname': 'First Name',
            'first name': 'First Name',
            'givenname': 'First Name',
            'lastname': 'Last Name',
            'last name': 'Last Name',
            'surname': 'Last Name',
            'nominaldate': 'Date',
            'date': 'Date',
            'workdate': 'Date',
            'day': 'Day',
            'start': 'Start',
            'starttime': 'Start',
            'earliest': 'Start',
            'stop': 'Stop',
            'end': 'Stop',
            'endtime': 'Stop',
            'latest': 'Stop',
            'supervisor': 'Supervisor',
            'queue': 'Queue',
            'shift': 'Shift',
            'batch': 'Batch',
            'code': 'Code',
            'status': 'Code',
            'reason': 'Code',
            'startdate': 'Start Date',
            'fromdate': 'Start Date',
            'stopdate': 'Stop Date',
            'enddate': 'Stop Date',
            'todate': 'Stop Date',
            'schedulestart': 'Schedule Start',
            'plannedstart': 'Schedule Start',
            'scheduledstart': 'Schedule Start'
        }

        for key, value in column_mapping.items():
            if key in name:
                return value

        if '-' in name:
            return name.split('-')[-1].strip().title()
        if ' ' in name:
            return name.split()[-1].strip().title()
        return name.title()

    def standardize_time_format(self, time_str: str) -> str:
        """Convert time to HH:MM format"""
        if pd.isna(time_str) or str(time_str).strip().upper() == 'OFF':
            return 'OFF'
        time_str = str(time_str).strip()
        if '1900-01-01 00:00:00' in time_str:
            return '00:00'
        if time_str == '00:00:00':
            return '00:00'
        if len(time_str) == 19 and time_str[4] == '-' and time_str[10] == ' ':
            return time_str[11:16]
        if ' ' in time_str and ':' in time_str:
            time_part = time_str.split()[-1]
            if time_part.count(':') == 2:
                return time_part[:5]
            return time_part[:5] if len(time_part) >= 5 else time_part
        if time_str.count(':') == 2:
            return time_str[:5]
        if ':' in time_str and len(time_str) < 5:
            parts = time_str.split(':')
            return f"{parts[0].zfill(2)}:{parts[1].zfill(2)}"
        return time_str[:5] if len(time_str) >= 5 else time_str

    def is_regular_shift(self, start_time: str, stop_time: str) -> bool:
        """Check if shift matches regular patterns"""
        if start_time == 'OFF' or stop_time == 'OFF':
            return True
        if start_time == '16:00' and stop_time == '00:00':
            return True
        return any(start_time == s and stop_time == e for s, e in self.regular_shifts)

    def get_shift_stop_time(self, start_time: str) -> str:
        """Get stop time for a given start time"""
        for s, e in self.regular_shifts:
            if start_time == s:
                return e
        return '00:00'

    def check_schedule_status(self, start_time: str, stop_time: str) -> str:
        """Determine schedule status"""
        if start_time == 'OFF' or stop_time == 'OFF':
            return ''
        if not self.is_regular_shift(start_time, stop_time):
            if self.auto_correct:
                if start_time == '00:00':
                    return 'Vacation'
                for s, e in self.regular_shifts:
                    if start_time == s:
                        return ''
            return 'Error'
        return ''

    def process_schedule_file(self, file_path: str) -> pd.DataFrame:
        """Process the main schedule file with prefixed columns"""
        logger.info(f"Processing schedule file: {file_path}")

        # Read file
        df = pd.read_excel(file_path) if file_path.endswith('.xlsx') else pd.read_csv(file_path)

        # Print original columns for debugging
        logger.info(f"Original columns: {list(df.columns)}")
        logger.info(f"Original shape: {df.shape}")

        # Standardize column names
        df.columns = [self.extract_column_name(col) for col in df.columns]
        logger.info(f"Standardized columns: {list(df.columns)}")

        # Validate and handle required columns
        required = ['ID', 'First Name', 'Last Name', 'Date', 'Start', 'Stop']
        available = [col for col in required if col in df.columns]

        # Handle name columns
        if 'Name' in df.columns and ('First Name' not in df.columns or 'Last Name' not in df.columns):
            names = df['Name'].str.split(n=1, expand=True)
            df['First Name'] = names[0]
            df['Last Name'] = names[1] if len(names.columns) > 1 else ''
            available.extend(['First Name', 'Last Name'])

        # Handle date column
        if 'Date' not in df.columns:
            date_cols = [col for col in df.columns if 'date' in col.lower()]
            if date_cols:
                df['Date'] = df[date_cols[0]]
                logger.info(f"Using {date_cols[0]} as Date column")
            else:
                raise ValueError("Required column 'Date' not found in data")

        # Process data
        if 'First Name' in df.columns and 'Last Name' in df.columns:
            df['Name'] = df['First Name'] + ' ' + df['Last Name']
            df = df.drop(columns=['First Name', 'Last Name'], errors='ignore')
        elif 'Name' not in df.columns:
            raise ValueError("Could not determine agent name from columns")

        # Clean and standardize IDs
        df['ID'] = df['ID'].apply(self.clean_id)

        # Format dates and times
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce').dt.strftime('%Y-%m-%d')

        for col in ['Start', 'Stop']:
            if col in df.columns:
                df[col] = df[col].fillna('OFF').apply(self.standardize_time_format)

        # Add status column
        df['Status'] = df.apply(lambda row: self.check_schedule_status(row['Start'], row['Stop']), axis=1)

        logger.info(f"Processed {len(df)} schedule records")
        return df

    def process_roster_file(self, file_path: str) -> pd.DataFrame:
        """Process the roster file with ID matching"""
        logger.info(f"Processing roster file: {file_path}")

        # Read file
        df = pd.read_excel(file_path) if file_path.endswith('.xlsx') else pd.read_csv(file_path)

        # Standardize column names
        df.columns = [self.extract_column_name(col) for col in df.columns]

        # Validate required columns
        required = ['ID', 'Queue', 'Supervisor', 'Shift', 'Batch']
        missing = [col for col in required if col not in df.columns]
        if missing:
            raise ValueError(f"Missing required columns in roster file: {', '.join(missing)}")

        # Clean and standardize IDs
        df['ID'] = df['ID'].apply(self.clean_id)

        logger.info(f"Processed {len(df)} roster records")
        return df

    def process_code_file(self, file_path: str) -> pd.DataFrame:
        """Process the code file with special statuses"""
        logger.info(f"Processing code file: {file_path}")

        # Read file
        df = pd.read_excel(file_path) if file_path.endswith('.xlsx') else pd.read_csv(file_path)

        # Find required columns
        id_col = next((col for col in df.columns if 'ID' in col.lower()), None)
        code_col = next((col for col in df.columns if any(x in col.lower() for x in ['code', 'status', 'reason'])), None)
        start_date_col = next((col for col in df.columns if any(x in col.lower() for x in ['start date', 'from date', 'start'])), None)
        stop_date_col = next((col for col in df.columns if any(x in col.lower() for x in ['stop date', 'end date', 'to date', 'stop'])), None)
        schedule_start_col = next((col for col in df.columns if any(x in col.lower() for x in ['schedule start', 'planned start', 'scheduled start'])), None)

        # Rename columns to standard names
        column_mapping = {}
        if id_col: column_mapping[id_col] = 'ID'
        if code_col: column_mapping[code_col] = 'Code'
        if start_date_col: column_mapping[start_date_col] = 'Start Date'
        if stop_date_col: column_mapping[stop_date_col] = 'Stop Date'
        if schedule_start_col: column_mapping[schedule_start_col] = 'Schedule Start'

        df = df.rename(columns=column_mapping)

        # Validate we have the minimum required columns
        required = ['ID', 'Start Date', 'Stop Date']
        missing = [col for col in required if col not in df.columns]
        if missing:
            raise ValueError(f"Missing required columns in code file: {', '.join(missing)}. Found columns: {list(df.columns)}")

        # Clean and standardize IDs
        df['ID'] = df['ID'].apply(self.clean_id)

        # Process dates and times
        df['Start Date'] = pd.to_datetime(df['Start Date'], errors='coerce')
        df['Stop Date'] = pd.to_datetime(df['Stop Date'], errors='coerce')
        df['Schedule Start'] = df['Schedule Start'].apply(self.standardize_time_format)

        # Expand date ranges
        code_entries = []
        for _, row in df.iterrows():
            if pd.notna(row['Start Date']) and pd.notna(row['Stop Date']):
                date_range = pd.date_range(row['Start Date'], row['Stop Date'])
                for date in date_range:
                    code_entries.append({
                        'ID': row['ID'],
                        'Date': date.strftime('%Y-%m-%d'),
                        'Code': row['Code'] if pd.notna(row['Code']) else '',
                        'Schedule Start': row['Schedule Start'] if pd.notna(row['Schedule Start']) else ''
                    })

        logger.info(f"Processed {len(code_entries)} code records")
        return pd.DataFrame(code_entries)

    def merge_data(self, df_schedule: pd.DataFrame, df_roster: pd.DataFrame, df_code: pd.DataFrame) -> pd.DataFrame:
        """Merge all data sources with robust ID matching"""
        logger.info("Merging data from all sources")

        # Create ID mapping dictionaries
        roster_dict = df_roster.set_index('ID').to_dict('index')
        code_dict = df_code.groupby(['ID', 'Date']).first().to_dict('index')

        # Get all unique IDs from roster for matching
        roster_ids = df_roster['ID'].unique().tolist()

        # Process each schedule record
        merged_data = []
        for _, row in df_schedule.iterrows():
            schedule_id = row['ID']
            schedule_date = row['Date']

            # Find matching roster data
            roster_match = None
            if schedule_id in roster_dict:
                roster_match = roster_dict[schedule_id]
            else:
                # Try fuzzy matching if exact match not found
                best_match = self.find_best_match(schedule_id, roster_ids)
                if best_match:
                    logger.warning(f"Using fuzzy match for ID {schedule_id} -> {best_match}")
                    roster_match = roster_dict[best_match]

            # Find matching code data
            code_match = None
            code_key = (schedule_id, schedule_date)
            if code_key in code_dict:
                code_match = code_dict[code_key]
            else:
                # Try fuzzy matching for code data
                for (id_key, date_key), values in code_dict.items():
                    if date_key == schedule_date and self.find_best_match(schedule_id, [id_key]):
                        code_match = values
                        logger.warning(f"Using fuzzy match for code data: {schedule_id} -> {id_key}")
                        break

            # Create merged record
            merged_record = {
                'ID': schedule_id,
                'Name': row['Name'],
                'Date': schedule_date,
                'Start': row['Start'],
                'Stop': row['Stop'],
                'Status': row['Status'],
                'Queue': roster_match['Queue'] if roster_match else '',
                'Supervisor': roster_match['Supervisor'] if roster_match else '',
                'Batch': roster_match['Batch'] if roster_match else '',
                'Shift': roster_match['Shift'] if roster_match else ''
            }

            # Apply code data if available
            if code_match:
                if code_match['Code']:
                    if merged_record['Status'] == 'Error':
                        merged_record['Status'] = f"{code_match['Code']}, Error"
                    elif merged_record['Status'] == 'Vacation':
                        merged_record['Status'] = f"{code_match['Code']}, Vacation"
                    else:
                        merged_record['Status'] = code_match['Code']

                if code_match['Schedule Start']:
                    merged_record['Start'] = code_match['Schedule Start']
                    if code_match['Schedule Start'] != 'OFF':
                        merged_record['Stop'] = self.get_shift_stop_time(code_match['Schedule Start'])

                    # Re-check status after applying code data
                    merged_record['Status'] = self.check_schedule_status(
                        merged_record['Start'],
                        merged_record['Stop']
                    ) if not merged_record['Status'] else merged_record['Status']

            merged_data.append(merged_record)

        # Create DataFrame from merged data
        df_merged = pd.DataFrame(merged_data)

        # Reorder columns
        final_columns = ['ID', 'Name', 'Date', 'Start', 'Stop',
                        'Queue', 'Supervisor', 'Batch', 'Shift', 'Status']
        df_merged = df_merged[final_columns]

        logger.info(f"Merged data contains {len(df_merged)} records")
        return df_merged

    def format_excel_output(self, df: pd.DataFrame, output_path: str):
        """Format the Excel output with tables and styles"""
        logger.info(f"Formatting Excel output: {output_path}")

        # Create a new workbook
        wb = load_workbook(output_path)
        ws = wb.active

        # Apply table formatting
        tab = Table(displayName="ScheduleTable", ref=ws.dimensions)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                             showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)

        # Apply column formatting
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

        # Apply header formatting
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        # Apply cell formatting
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                if cell.column_letter in ['D', 'E']:  # Start and Stop columns
                    cell.alignment = Alignment(horizontal='center')
                if cell.column_letter == 'J' and 'Error' in str(cell.value):  # Status column
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Save the formatted workbook
        wb.save(output_path)
        logger.info("Excel formatting completed")

    def process(self, schedule_file: str, roster_file: str, code_file: str) -> str:
        """Main processing method"""
        try:
            # Process each file
            df_schedule = self.process_schedule_file(schedule_file)
            df_roster = self.process_roster_file(roster_file)
            df_code = self.process_code_file(code_file)

            # Merge all data
            df_merged = self.merge_data(df_schedule, df_roster, df_code)

            # Generate output filename
            file_name = os.path.splitext(schedule_file)[0]
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"{file_name}_{timestamp}_clean.xlsx"

            # Save to Excel
            df_merged.to_excel(output_file, index=False)

            # Format the Excel output
            self.format_excel_output(df_merged, output_file)

            # Show statistics
            self.show_statistics(df_merged)

            return output_file

        except Exception as e:
            logger.error(f"Error processing files: {e}")
            raise

    def show_statistics(self, df: pd.DataFrame):
        """Show processing statistics"""
        if 'Status' not in df.columns:
            return

        error_count = df['Status'].str.contains('Error').sum()
        vac_count = df['Status'].str.contains('Vacation').sum()
        code_count = df['Status'].str.contains('[A-Za-z]').sum() - error_count - vac_count

        logger.info("\nProcessing Statistics:")
        logger.info(f"Total records: {len(df)}")
        logger.info(f"Records with shift errors: {error_count}")
        logger.info(f"Records marked as Vacation: {vac_count}")
        logger.info(f"Records with special codes: {code_count}")

        if error_count > 0:
            logger.info("\nSample of schedules with shift errors:")
            error_schedules = df[df['Status'].str.contains('Error')]
            for _, row in error_schedules.head(5).iterrows():
                logger.info(f"  {row['Name']} - {row['Date']}: {row['Start']}-{row['Stop']} ({row['Status']})")
            if len(error_schedules) > 5:
                logger.info(f"  ... and {len(error_schedules) - 5} more")

def main():
    import argparse
    parser = argparse.ArgumentParser(description='Process schedule data from three input files.')
    parser.add_argument('schedule_file', help='Path to the schedule file (Excel or CSV)')
    parser.add_argument('roster_file', help='Path to the roster file (Excel or CSV)')
    parser.add_argument('code_file', help='Path to the code file (Excel or CSV)')
    parser.add_argument('--auto-correct', action='store_true',
                       help='Automatically correct shift errors using predefined shifts')
    args = parser.parse_args()

    # Check if files exist
    for file_path in [args.schedule_file, args.roster_file, args.code_file]:
        if not os.path.exists(file_path):
            logger.error(f"File not found: {file_path}")
            sys.exit(1)

    try:
        processor = ScheduleProcessor(auto_correct=args.auto_correct)
        output_file = processor.process(
            args.schedule_file,
            args.roster_file,
            args.code_file
        )
        logger.info(f"Processing completed successfully! Output: {output_file}")
    except Exception as e:
        logger.error(f"Processing failed: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()