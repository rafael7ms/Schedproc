import openpyxl

wb = openpyxl.load_workbook('Agent Schedules (Dec 1st - 21st)_20251218_151010_seating_arrangement.xlsx')
print('Sheets in workbook:', wb.sheetnames)
print()

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'{sheet_name}: {ws.max_row} rows x {ws.max_column} cols')

print()
print('Legend sheet has statistics:')
legend_ws = wb['Legend']
for row in range(22, 27):
    print(f'  {legend_ws[f"A{row}"].value}')
