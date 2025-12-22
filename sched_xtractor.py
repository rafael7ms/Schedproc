import pandas as pd
import os
import sys
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

def process_schedule_data(input_file):
    """
    Process schedule data to ensure proper format before creating report
    """
    try:
        # Read the input file
        if input_file.endswith('.xlsx'):
            df = pd.read_excel(input_file)
        elif input_file.endswith('.csv'):
            df = pd.read_csv(input_file)
        else:
            raise ValueError("File must be .xlsx or .csv format")

        print(f"Processing schedule data from: {input_file}")

        # Check if we need to create Name column from First Name and Last Name
        if 'First Name' in df.columns and 'Last Name' in df.columns and 'Name' not in df.columns:
            df['Name'] = df['First Name'] + ' ' + df['Last Name']
            print("Created Name column from First Name and Last Name")

        # Ensure we have required columns
        required_columns = ['ID', 'Name', 'Date', 'Start', 'Stop', 'Queue', 'Supervisor', 'Batch', 'Shift', 'Status']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            print(f"Warning: Missing columns {missing_columns}. Processing with available data.")

        return df

    except Exception as e:
        print(f"Error processing schedule data: {e}")
        raise

def create_agent_schedule_table(input_file):
    """
    Create a schedule table with ID, Full Name, Queue, and date columns showing schedule information
    Output file will end with _agent_schedules.xlsx
    """
    try:
        # Read the input file
        if input_file.endswith('.xlsx'):
            df = pd.read_excel(input_file)
        elif input_file.endswith('.csv'):
            df = pd.read_csv(input_file)
        else:
            raise ValueError("File must be .xlsx or .csv format")

        print(f"Processing file: {input_file}")
        print(f"Original shape: {df.shape}")
        print(f"Columns: {list(df.columns)}")

        # Ensure Date column is datetime
        df['Date'] = pd.to_datetime(df['Date'])

        # Create the shift information
        def create_shift_info(row):
            start = str(row['Start']).strip().upper() if pd.notna(row['Start']) else ''
            stop = str(row['Stop']).strip().upper() if pd.notna(row['Stop']) else ''
            status = str(row['Status']).strip().upper() if pd.notna(row['Status']) else ''

            # If status is not empty and not OFF, return the status
            if status and status != 'OFF':
                return status
            # If either start or stop is OFF, return OFF
            elif start == 'OFF' or stop == 'OFF' or pd.isna(row['Start']) or pd.isna(row['Stop']):
                return 'OFF'
            else:
                return f"{start}-{stop}"

        df['Shift_Info'] = df.apply(create_shift_info, axis=1)

        # Get unique agents and dates
        agents = df[['ID', 'Name', 'Queue', 'Supervisor', 'Batch', 'Shift']].drop_duplicates()
        dates = sorted(df['Date'].unique())

        # Create a result DataFrame with all agents
        result_df = agents.copy()

        # Add a column for each date
        for date in dates:
            date_str = date.strftime('%Y-%m-%d')
            # Get shift info for this date
            date_data = df[df['Date'] == date]
            # Merge with result_df to add the shift info
            result_df = result_df.merge(
                date_data[['ID', 'Shift_Info']],
                on='ID',
                how='left'
            ).rename(columns={'Shift_Info': date_str})

        # Fill NaN values with 'OFF'
        date_columns = [date.strftime('%Y-%m-%d') for date in dates]
        result_df[date_columns] = result_df[date_columns].fillna('OFF')

        # Create output filename
        file_name_without_ext = os.path.splitext(os.path.basename(input_file))[0]
        output_file = f"{file_name_without_ext}_agent_schedules.xlsx"

        # Define queue colors (add more as needed)
        queue_colors = {
            'Sales': 'FFC6EFCE',      # Light green
            'Support': 'FFDDEBF7',    # Light blue
            'Billing': 'FFFFE699',    # Light yellow
            'Retention': 'FFFDE9D9',  # Light orange
            'Escalations': 'FFFFC7CE', # Light red
            'Default': 'FFFFFFFF'     # White
        }

        # Define status fills
        status_fills = {
            'VACATION': PatternFill(start_color='FFFFC7CE', end_color='FFFFC7CE', fill_type='solid'),  # Light red
            'PTO': PatternFill(start_color='FFFFC7CE', end_color='FFFFC7CE', fill_type='solid'),      # Light red
            'TRAINING': PatternFill(start_color='FFC6EFCE', end_color='FFC6EFCE', fill_type='solid'), # Light green
            'NESTING': PatternFill(start_color='FFDDEBF7', end_color='FFDDEBF7', fill_type='solid'),  # Light blue
            'OFF': PatternFill(start_color='FFD3D3D3', end_color='FFD3D3D3', fill_type='solid')      # Light gray
        }

        # Save to Excel with formatting using openpyxl
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            result_df.to_excel(writer, index=False, sheet_name='Schedule')

            worksheet = writer.sheets['Schedule']

            # Define styles
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Apply formatting to all cells
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                    # Make header row bold
                    if cell.row == 1:
                        cell.font = Font(bold=True)
                        continue

                    # Get the cell value
                    cell_value = str(cell.value).strip().upper() if cell.value else ''

                    # Apply status-based formatting for date columns
                    if cell.column > 6:  # Date columns start after the first 6 columns
                        # Apply status color if applicable
                        for status, fill in status_fills.items():
                            if status in cell_value:
                                cell.fill = fill
                                break
                        else:
                            # If no status match, apply queue-based coloring
                            queue = str(result_df.iloc[cell.row-2]['Queue']).strip()
                            color_code = queue_colors.get(queue, queue_colors['Default'])
                            cell.fill = PatternFill(start_color=color_code, end_color=color_code, fill_type='solid')

                    # Apply queue-based coloring to the Queue column (column 3)
                    if cell.column == 3:
                        queue = str(cell.value).strip()
                        color_code = queue_colors.get(queue, queue_colors['Default'])
                        cell.fill = PatternFill(start_color=color_code, end_color=color_code, fill_type='solid')

            # Auto-adjust column widths
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter  # Get the column name
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)  # Cap at 30 for readability
                worksheet.column_dimensions[column].width = adjusted_width

        print(f"\nSchedule report completed successfully!")
        print(f"Output saved to: {output_file}")
        print(f"Final shape: {result_df.shape}")
        print(f"Columns: {list(result_df.columns)}")

        return result_df, output_file

    except Exception as e:
        print(f"Error processing file: {e}")
        raise

def main():
    # Check if file path is provided as command line argument
    if len(sys.argv) != 2:
        print("Usage: python agent_schedule_generator.py <input_file_path>")
        print("Example: python agent_schedule_generator.py 'Schedule data.xlsx'")
        sys.exit(1)

    input_file = sys.argv[1]

    # Check if file exists
    if not os.path.exists(input_file):
        print(f"Error: File '{input_file}' not found.")
        sys.exit(1)

    try:
        # Process the data to ensure proper format
        processed_df = process_schedule_data(input_file)

        # Create the agent schedule report
        df_report, output_file = create_agent_schedule_table(input_file)

    except Exception as e:
        print(f"Process failed: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()