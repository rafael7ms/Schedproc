import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime, time
import numpy as np
import argparse
import os
import sys

def load_data(file_path):
    """Load the input data and add Seat column"""
    try:
        df = pd.read_excel(file_path)
        df['Seat'] = np.nan  # Initialize Seat column
        return df
    except Exception as e:
        print(f"Error reading input file: {e}")
        sys.exit(1)

def parse_time(time_val):
    """Parse time value, handling different formats"""
    if isinstance(time_val, str):
        if time_val.upper() == 'OFF':
            return None
        try:
            return datetime.strptime(time_val, '%H:%M').time()
        except ValueError:
            try:
                return datetime.strptime(time_val, '%H:%M:%S').time()
            except ValueError:
                return None
    elif isinstance(time_val, datetime):
        return time_val.time()
    elif isinstance(time_val, time):
        return time_val
    return None

def should_ignore_agent(row):
    """Determine if an agent should be ignored for seating"""
    # Ignore agents with Vacation or Training status
    if row['Status'] in ['Vacation', 'Training']:
        return True
    # Ignore agents with Start = "OFF"
    if isinstance(row['Start'], str) and row['Start'].upper() == 'OFF':
        return True
    return False

def get_subarea_combinations():
    """Return possible subarea combinations for nesting agents"""
    return [
        ['OPS1-A'],  # 12 seats
        ['OPS1-B'],  # 6 seats
        ['OPS1-C'],  # 6 seats
        ['OPS1-D'],  # 15 seats
        ['OPS1-E'],  # 6 seats
        ['OPS1-A', 'OPS1-B'],  # 18 seats
        ['OPS1-B', 'OPS1-C'],  # 12 seats
        ['OPS1-C', 'OPS1-D'],  # 21 seats
        ['OPS1-C', 'OPS1-E'],  # 12 seats
    ]

def create_seat_structure():
    """Create the seat structure with all metadata"""
    seats = []

    # OPS1 - Seats 1-47 (46,47 reserved)
    for seat in range(1, 48):
        if seat in [46, 47]:
            reserved = True
            subarea = None
        else:
            reserved = False
            if seat <= 12:
                subarea = 'OPS1-A'
            elif seat <= 18:
                subarea = 'OPS1-B'
            elif seat <= 24:
                subarea = 'OPS1-C'
            elif seat <= 39:
                subarea = 'OPS1-D'
            else:
                subarea = 'OPS1-E'

        seats.append({
            'Seat': seat,
            'Area': 'OPS1',
            'Subarea': subarea,
            'Reserved': reserved,
            'Queue': None
        })

    # OPS2 - Seats 48-61 (61 reserved)
    for seat in range(48, 62):
        seats.append({
            'Seat': seat,
            'Area': 'OPS2',
            'Subarea': None,
            'Reserved': seat == 61,
            'Queue': 'IBC Support' if not (seat == 61) else None
        })

    # OPS3 - Seats 62-69
    for seat in range(62, 70):
        seats.append({
            'Seat': seat,
            'Area': 'OPS3',
            'Subarea': None,
            'Reserved': False,
            'Queue': 'IBC Support'
        })

    return pd.DataFrame(seats)

def assign_seats(df):
    """Assign seats to all agents based on the rules"""
    # Create seat structure
    seat_df = create_seat_structure()
    
    # Add Assigned column to track which seats are taken
    seat_df['Assigned'] = False
    
    # Get unique dates
    dates = df['Date'].unique()
    
    for date in dates:
        # Reset assigned status for each date
        seat_df['Assigned'] = False
        
        # Get agents for this date
        date_agents = df[df['Date'] == date]
        
        # Separate agents by type
        nesting_agents = []
        ibc_agents = []
        other_agents = []
        
        for _, agent in date_agents.iterrows():
            if should_ignore_agent(agent):
                continue
            
            # Parse start and stop times
            start_time = parse_time(agent['Start'])
            stop_time = parse_time(agent['Stop'])
            
            if start_time is None or stop_time is None:
                continue
                
            agent_dict = {
                'Index': agent.name,  # Keep track of original row index
                'Name': agent['Name'],
                'Queue': agent['Queue'],
                'Start': start_time,
                'Stop': stop_time,
                'Status': agent['Status']
            }
            
            if agent['Status'] == 'Nesting':
                nesting_agents.append(agent_dict)
            elif agent['Queue'] == 'IBC Support':
                ibc_agents.append(agent_dict)
            elif agent['Queue'] in ['Customer Support', 'BNS']:
                other_agents.append(agent_dict)
        
        # Assign nesting agents first
        assign_nesting_agents(nesting_agents, seat_df, df)
        
        # Assign IBC agents
        assign_ibc_agents(ibc_agents, seat_df, df)
        
        # Assign other agents
        assign_other_agents(other_agents, seat_df, df)
    
    return df

def assign_nesting_agents(agents, seat_df, df):
    """Assign seats to nesting agents"""
    if not agents:
        return
    
    # Get possible subarea combinations
    combinations = get_subarea_combinations()
    
    # Find the smallest combination that can fit all nesting agents
    selected_combination = None
    for combo in combinations:
        total_seats = sum(len(seat_df[(seat_df['Subarea'] == subarea) & (~seat_df['Reserved']) & (~seat_df['Assigned'])]) for subarea in combo)
        if total_seats >= len(agents):
            selected_combination = combo
            break
    
    if not selected_combination:
        # If no combination can fit all, use the largest one
        selected_combination = max(combinations, key=lambda combo:
            sum(len(seat_df[(seat_df['Subarea'] == subarea) & (~seat_df['Reserved']) & (~seat_df['Assigned'])]) for subarea in combo))
    
    # Get available seats in the selected combination
    available_seats = seat_df[
        (seat_df['Subarea'].isin(selected_combination)) &
        (~seat_df['Reserved']) &
        (~seat_df['Assigned'])
    ]['Seat'].tolist()
    
    # Assign agents to seats
    for i, agent in enumerate(agents):
        if i < len(available_seats):
            seat_num = available_seats[i]
            df.at[agent['Index'], 'Seat'] = seat_num
            # Mark seat as taken
            seat_df.loc[seat_df['Seat'] == seat_num, 'Assigned'] = True

def assign_ibc_agents(agents, seat_df, df):
    """Assign seats to IBC Support agents"""
    if not agents:
        return
    
    # Count IBC agents
    count = len(agents)
    
    # Rule: If more than 13 IBC agents, at least 2 should be in OPS3 same shift
    if count > 13:
        # Group agents by shift
        shift_groups = {}
        for agent in agents:
            shift_key = (agent['Start'], agent['Stop'])
            if shift_key not in shift_groups:
                shift_groups[shift_key] = []
            shift_groups[shift_key].append(agent)
        
        # Get shifts with most agents
        sorted_shifts = sorted(shift_groups.items(), key=lambda x: len(x[1]), reverse=True)
        top_shifts = [shift[0] for shift in sorted_shifts[:2]]
        
        # Assign 2 agents from the same shift to OPS3
        ops3_seats = seat_df[
            (seat_df['Area'] == 'OPS3') &
            (~seat_df['Reserved']) &
            (~seat_df['Assigned'])
        ]['Seat'].tolist()
        
        assigned = 0
        for shift in top_shifts:
            for agent in shift_groups[shift]:
                if assigned >= 2:
                    break
                if ops3_seats:
                    seat_num = ops3_seats.pop(0)
                    df.at[agent['Index'], 'Seat'] = seat_num
                    # Mark seat as taken
                    seat_df.loc[seat_df['Seat'] == seat_num, 'Assigned'] = True
                    assigned += 1
    
    # Assign remaining IBC agents to OPS2 and OPS3
    ibc_seats = seat_df[
        (seat_df['Queue'] == 'IBC Support') &
        (~seat_df['Reserved']) &
        (~seat_df['Assigned'])
    ]['Seat'].tolist()
    
    # Sort agents by shift to group them together
    agents_sorted = sorted(agents, key=lambda x: (x['Start'], x['Stop']))
    
    for agent in agents_sorted:
        if not ibc_seats:
            break
        seat_num = ibc_seats.pop(0)
        df.at[agent['Index'], 'Seat'] = seat_num
        # Mark seat as taken
        seat_df.loc[seat_df['Seat'] == seat_num, 'Assigned'] = True

def assign_other_agents(agents, seat_df, df):
    """Assign seats to Customer Support and BNS agents"""
    if not agents:
        return
    
    # Sort agents by queue and shift to group them together
    agents_sorted = sorted(agents, key=lambda x: (x['Queue'], x['Start'], x['Stop']))
    
    # Get available seats in OPS1 (excluding reserved seats)
    available_seats = seat_df[
        (seat_df['Area'] == 'OPS1') &
        (~seat_df['Reserved']) &
        (~seat_df['Assigned'])
    ]['Seat'].tolist()
    
    for agent in agents_sorted:
        if not available_seats:
            break
        seat_num = available_seats.pop(0)
        df.at[agent['Index'], 'Seat'] = seat_num
        # Mark seat as taken
        seat_df.loc[seat_df['Seat'] == seat_num, 'Assigned'] = True

def create_reports(df, output_path):
    """Create the Excel reports from the DataFrame with assigned seats"""
    wb = Workbook()
    
    # Define color fills
    nesting_fill = PatternFill(start_color='FFFFCC99', end_color='FFFFCC99', fill_type='solid')
    bns_fill = PatternFill(start_color='FFCCFFCC', end_color='FFCCFFCC', fill_type='solid')
    customer_fill = PatternFill(start_color='FFCCE5FF', end_color='FFCCE5FF', fill_type='solid')
    ibc_fill = PatternFill(start_color='FFFFCCCC', end_color='FFFFCCCC', fill_type='solid')
    reserved_fill = PatternFill(start_color='FFD3D3D3', end_color='FFD3D3D3', fill_type='solid')
    empty_fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
    dual_shift_fill = PatternFill(start_color='FFE6F2FF', end_color='FFE6F2FF', fill_type='solid')  # Light blue for dual shifts
    
    # Get unique dates
    dates = pd.to_datetime(df['Date']).dt.date.unique()
    
    # Create seat structure for reference
    seat_df = create_seat_structure()
    
    # Morning Shifts sheet
    ws_morning = wb.active
    ws_morning.title = "Morning Shifts"
    
    # Create header for morning shifts
    morning_header = ['Area', 'Subarea', 'Seat']
    for date in sorted(dates):
        morning_header.append(date.strftime('%Y-%m-%d'))
    ws_morning.append(morning_header)
    
    # Add data for morning shifts - include ALL seats
    for _, seat_row in seat_df.iterrows():
        data_row = [seat_row['Area'], seat_row['Subarea'], seat_row['Seat']]
        
        for date in sorted(dates):
            # Find agent assigned to this seat on this date during morning shift
            agent = df[(pd.to_datetime(df['Date']).dt.date == date) & 
                      (df['Seat'] == seat_row['Seat'])]
            
            # Check if agent works in morning (start before 12:00)
            if not agent.empty:
                start_time = parse_time(agent.iloc[0]['Start'])
                if start_time and start_time < time(12, 0):
                    agent_name = agent.iloc[0]['Name']
                    data_row.append(agent_name)
                    continue
            
            # For reserved seats
            if seat_row['Reserved']:
                data_row.append("RESERVED")
            else:
                data_row.append("EMPTY")
        
        ws_morning.append(data_row)
    
    # Apply formatting to morning sheet
    for row_idx, row in enumerate(ws_morning.iter_rows(min_row=2, max_row=ws_morning.max_row, min_col=1, max_col=ws_morning.max_column), 2):
        for col_idx, cell in enumerate(row, 1):
            if col_idx <= 3:  # Skip Area, Subarea, Seat columns
                continue
                
            # Get the seat number from column C of this row
            seat_num = ws_morning.cell(row=row_idx, column=3).value
            
            if seat_num:
                # Get the date from the header
                date_col = col_idx - 3  # Column index relative to dates
                if date_col <= len(dates):
                    date = sorted(dates)[date_col-1] if date_col <= len(dates) else None
                    
                    if date:
                        # Find agent assigned to this seat on this date
                        agent = df[(pd.to_datetime(df['Date']).dt.date == date) & (df['Seat'] == seat_num)]
                        
                        if not agent.empty:
                            status = agent.iloc[0]['Status']
                            queue = agent.iloc[0]['Queue']
                            
                            if seat_df[seat_df['Seat'] == seat_num]['Reserved'].iloc[0]:
                                cell.fill = reserved_fill
                            elif status == 'Nesting':
                                cell.fill = nesting_fill
                            elif queue == 'BNS':
                                cell.fill = bns_fill
                            elif queue == 'Customer Support':
                                cell.fill = customer_fill
                            elif queue == 'IBC Support':
                                cell.fill = ibc_fill
                        elif cell.value == "EMPTY":
                            cell.fill = empty_fill
                        elif cell.value == "RESERVED":
                            cell.fill = reserved_fill
    
    # Create table for morning shifts
    if ws_morning.max_row > 1:  # Only create table if there's data
        tab = Table(displayName="MorningTable", ref=f"A1:{chr(64 + len(dates) + 3)}{ws_morning.max_row}")
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                              showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws_morning.add_table(tab)
    
    # Auto-fit columns for morning shifts
    for column in ws_morning.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws_morning.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # Afternoon Shifts sheet
    ws_afternoon = wb.create_sheet("Afternoon Shifts")
    
    # Create header for afternoon shifts
    afternoon_header = ['Area', 'Subarea', 'Seat']
    for date in sorted(dates):
        afternoon_header.append(date.strftime('%Y-%m-%d'))
    ws_afternoon.append(afternoon_header)
    
    # Add data for afternoon shifts - include ALL seats
    for _, seat_row in seat_df.iterrows():
        data_row = [seat_row['Area'], seat_row['Subarea'], seat_row['Seat']]
        
        for date in sorted(dates):
            # Find agent assigned to this seat on this date during afternoon shift
            agent = df[(pd.to_datetime(df['Date']).dt.date == date) & 
                      (df['Seat'] == seat_row['Seat'])]
            
            # Check if agent works in afternoon (start at or after 12:00)
            if not agent.empty:
                start_time = parse_time(agent.iloc[0]['Start'])
                if start_time and start_time >= time(12, 0):
                    agent_name = agent.iloc[0]['Name']
                    data_row.append(agent_name)
                    continue
            
            # For reserved seats
            if seat_row['Reserved']:
                data_row.append("RESERVED")
            else:
                data_row.append("EMPTY")
        
        ws_afternoon.append(data_row)
    
    # Apply formatting to afternoon sheet
    for row_idx, row in enumerate(ws_afternoon.iter_rows(min_row=2, max_row=ws_afternoon.max_row, min_col=1, max_col=ws_afternoon.max_column), 2):
        for col_idx, cell in enumerate(row, 1):
            if col_idx <= 3:  # Skip Area, Subarea, Seat columns
                continue
                
            # Get the seat number from column C of this row
            seat_num = ws_afternoon.cell(row=row_idx, column=3).value
            
            if seat_num:
                # Get the date from the header
                date_col = col_idx - 3  # Column index relative to dates
                if date_col <= len(dates):
                    date = sorted(dates)[date_col-1] if date_col <= len(dates) else None
                    
                    if date:
                        # Find agent assigned to this seat on this date
                        agent = df[(pd.to_datetime(df['Date']).dt.date == date) & (df['Seat'] == seat_num)]
                        
                        if not agent.empty:
                            status = agent.iloc[0]['Status']
                            queue = agent.iloc[0]['Queue']
                            
                            if seat_df[seat_df['Seat'] == seat_num]['Reserved'].iloc[0]:
                                cell.fill = reserved_fill
                            elif status == 'Nesting':
                                cell.fill = nesting_fill
                            elif queue == 'BNS':
                                cell.fill = bns_fill
                            elif queue == 'Customer Support':
                                cell.fill = customer_fill
                            elif queue == 'IBC Support':
                                cell.fill = ibc_fill
                        elif cell.value == "EMPTY":
                            cell.fill = empty_fill
                        elif cell.value == "RESERVED":
                            cell.fill = reserved_fill
    
    # Create table for afternoon shifts
    if ws_afternoon.max_row > 1:  # Only create table if there's data
        tab = Table(displayName="AfternoonTable", ref=f"A1:{chr(64 + len(dates) + 3)}{ws_afternoon.max_row}")
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                              showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws_afternoon.add_table(tab)
    
    # Auto-fit columns for afternoon shifts
    for column in ws_afternoon.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws_afternoon.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # All Shifts sheet
    ws_all = wb.create_sheet("All Shifts")
    
    # Create header for all shifts
    all_header = ['Area', 'Subarea', 'Seat']
    for date in sorted(dates):
        all_header.append(date.strftime('%Y-%m-%d'))
    ws_all.append(all_header)
    
    # Add data for all shifts - include ALL seats
    for _, seat_row in seat_df.iterrows():
        data_row = [seat_row['Area'], seat_row['Subarea'], seat_row['Seat']]
        
        for date in sorted(dates):
            # Find all agents assigned to this seat on this date
            agents = df[(pd.to_datetime(df['Date']).dt.date == date) & 
                       (df['Seat'] == seat_row['Seat'])]
            
            if not agents.empty:
                # Check if there are multiple agents (morning and afternoon)
                morning_agent = None
                afternoon_agent = None
                
                for _, agent in agents.iterrows():
                    start_time = parse_time(agent['Start'])
                    if start_time:
                        if start_time < time(12, 0):
                            morning_agent = agent['Name']
                        else:
                            afternoon_agent = agent['Name']
                
                # Format the cell content
                if morning_agent and afternoon_agent:
                    # Both shifts - show both names
                    data_row.append(f"{morning_agent}/{afternoon_agent}")
                elif morning_agent:
                    # Only morning shift
                    data_row.append(morning_agent)
                elif afternoon_agent:
                    # Only afternoon shift
                    data_row.append(afternoon_agent)
                else:
                    # Fallback
                    agent_name = agents.iloc[0]['Name']
                    data_row.append(agent_name)
            elif seat_row['Reserved']:
                data_row.append("RESERVED")
            else:
                data_row.append("EMPTY")
        
        ws_all.append(data_row)
    
    # Apply formatting to all shifts sheet
    for row_idx, row in enumerate(ws_all.iter_rows(min_row=2, max_row=ws_all.max_row, min_col=1, max_col=ws_all.max_column), 2):
        for col_idx, cell in enumerate(row, 1):
            if col_idx <= 3:  # Skip Area, Subarea, Seat columns
                continue
                
            # Get the seat number from column C of this row
            seat_num = ws_all.cell(row=row_idx, column=3).value
            
            if seat_num:
                # Get the date from the header
                date_col = col_idx - 3  # Column index relative to dates
                if date_col <= len(dates):
                    date = sorted(dates)[date_col-1] if date_col <= len(dates) else None
                    
                    if date:
                        # Find all agents assigned to this seat on this date
                        agents = df[(pd.to_datetime(df['Date']).dt.date == date) & (df['Seat'] == seat_num)]
                        
                        if not agents.empty:
                            # Check if there are multiple agents (dual shift)
                            morning_count = 0
                            afternoon_count = 0
                            
                            for _, agent in agents.iterrows():
                                start_time = parse_time(agent['Start'])
                                if start_time:
                                    if start_time < time(12, 0):
                                        morning_count += 1
                                    else:
                                        afternoon_count += 1
                            
                            # Apply dual shift formatting if both shifts are present
                            if morning_count > 0 and afternoon_count > 0:
                                cell.fill = dual_shift_fill
                                cell.font = Font(bold=True)
                            else:
                                # Apply regular formatting
                                agent = agents.iloc[0]
                                status = agent['Status']
                                queue = agent['Queue']
                                
                                if seat_df[seat_df['Seat'] == seat_num]['Reserved'].iloc[0]:
                                    cell.fill = reserved_fill
                                elif status == 'Nesting':
                                    cell.fill = nesting_fill
                                elif queue == 'BNS':
                                    cell.fill = bns_fill
                                elif queue == 'Customer Support':
                                    cell.fill = customer_fill
                                elif queue == 'IBC Support':
                                    cell.fill = ibc_fill
                        elif cell.value == "EMPTY":
                            cell.fill = empty_fill
                        elif cell.value == "RESERVED":
                            cell.fill = reserved_fill
    
    # Create table for all shifts
    if ws_all.max_row > 1:  # Only create table if there's data
        tab = Table(displayName="AllTable", ref=f"A1:{chr(64 + len(dates) + 3)}{ws_all.max_row}")
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                              showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws_all.add_table(tab)
    
    # Auto-fit columns for all shifts
    for column in ws_all.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws_all.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # Save workbook
    try:
        wb.save(output_path)
        print(f"Seating arrangement successfully saved to {output_path}")
    except Exception as e:
        print(f"Error saving output file: {e}")
        sys.exit(1)

def generate_output_filename(input_filename):
    """Generate output filename based on input filename and current timestamp"""
    # Get the base name without extension
    base_name = os.path.splitext(os.path.basename(input_filename))[0]
    
    # Split at first underscore
    parts = base_name.split('_', 1)
    prefix = parts[0] if parts else base_name
    
    # Get current timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Create output filename
    output_filename = f"{prefix}_{timestamp}_seating_arrangement.xlsx"
    
    return output_filename

def main():
    # Set up argument parser
    parser = argparse.ArgumentParser(description='Generate seating arrangement for contact center agents.')
    parser.add_argument('input_file', help='Path to the input Excel file with agent schedules')
    
    # Parse arguments
    args = parser.parse_args()
    
    # Check if input file exists
    if not os.path.exists(args.input_file):
        print(f"Error: Input file '{args.input_file}' not found.")
        sys.exit(1)
    
    # Generate output filename
    output_file = generate_output_filename(args.input_file)
    
    # Load data
    print(f"Loading input file: {args.input_file}")
    df = load_data(args.input_file)
    
    # Assign seats
    print("Assigning seats...")
    df = assign_seats(df)
    
    # Create reports
    print(f"Generating output file: {output_file}")
    create_reports(df, output_file)

if __name__ == "__main__":
    main()
    