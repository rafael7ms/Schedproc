import argparse
import json
import logging
import sys
from datetime import datetime, time, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Union

import pandas as pd
import ollama
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler(sys.stdout)]
)
logger = logging.getLogger(__name__)

# Constants
SEAT_AREAS = {
    'OPS1-A': range(1, 13),
    'OPS1-B': range(13, 19),
    'OPS1-C': range(19, 25),
    'OPS1-D': range(25, 40),
    'OPS1-E': range(40, 48),
    'OPS2': range(48, 62),
    'OPS3': range(62, 70),
    'TRN': range(71, 101)
}

RESERVED_SEATS = [61]
TOTAL_SEATS = 69

SHIFT_TIMES = {
    '5AM': (time(5, 0), time(14, 0)),
    '6AM': (time(6, 0), time(15, 0)),
    '7AM': (time(7, 0), time(16, 0)),
    '10AM': (time(10, 0), time(19, 0)),
    '11AM': (time(11, 0), time(20, 0)),
    '2PM': (time(14, 0), time(22, 0)),
    '3PM': (time(15, 0), time(23, 0)),
    '3:30PM': (time(15, 30), time(23, 30)),
    '4PM': (time(16, 0), time(0, 0)),
    '5PM': (time(17, 0), time(1, 0))
}

class SeatAssignmentSystem:
    def __init__(self, ollama_host: str, model: str = "llama3"):
        self.ollama_host = ollama_host
        self.model = model
        self.client = ollama.Client(host=ollama_host)
        self.agents_df = None
        self.rules = ""
        self.date_range = None
        self.area_capacities = self._calculate_area_capacities()

    def _calculate_area_capacities(self) -> Dict[str, int]:
        """Calculate capacity for each area."""
        return {area: len(seats) for area, seats in SEAT_AREAS.items()}

    def validate_connection(self) -> bool:
        """Validate connection to Ollama server."""
        try:
            self.client.list()
            return True
        except Exception as e:
            logger.error(f"Failed to connect to Ollama server: {e}")
            return False

    def load_agent_data(self, file_path: str) -> bool:
        """Load agent data from Excel file."""
        try:
            self.agents_df = pd.read_excel(file_path)

            # Validate required columns
            required_columns = {'ID', 'Name', 'Date', 'Queue', 'Shift', 'Supervisor', 'Batch'}
            if not required_columns.issubset(self.agents_df.columns):
                missing = required_columns - set(self.agents_df.columns)
                raise ValueError(f"Missing required columns: {missing}")

            # Convert and validate dates
            self.agents_df['Date'] = pd.to_datetime(self.agents_df['Date'], errors='coerce')
            self.agents_df = self.agents_df.dropna(subset=['Date'])

            # Filter out invalid shifts
            self.agents_df = self.agents_df[
                ~self.agents_df['Shift'].str.upper().isin(['OFF', 'TRAINING', 'VACATION'])
            ]

            # Get date range
            self.date_range = pd.date_range(
                start=self.agents_df['Date'].min(),
                end=self.agents_df['Date'].max()
            )

            logger.info(f"Loaded agent data for {len(self.agents_df)} records")
            return True
        except Exception as e:
            logger.error(f"Error loading agent data: {e}")
            return False

    def load_rules(self, rules_input: Union[str, None]) -> bool:
        """Load rules from file or direct input."""
        if rules_input is None:
            self.rules = ""
            return True

        if Path(rules_input).exists():
            try:
                with open(rules_input, 'r') as f:
                    self.rules = f.read()
                logger.info("Loaded rules from file")
                return True
            except Exception as e:
                logger.error(f"Error reading rules file: {e}")
                return False
        else:
            self.rules = rules_input
            logger.info("Using rules from direct input")
            return True

    def _parse_shift_time(self, shift: str) -> Optional[Tuple[time, time]]:
        """Parse shift string into start and end times."""
        shift = shift.upper().strip()

        if shift in SHIFT_TIMES:
            return SHIFT_TIMES[shift]

        # Try to parse custom shift format
        time_pattern = r'^(\d{1,2}):(\d{2})-(\d{1,2}):(\d{2})$'
        match = re.match(time_pattern, shift)
        if match:
            start_h, start_m, end_h, end_m = map(int, match.groups())
            return (time(start_h, start_m), time(end_h, end_m))

        logger.warning(f"Could not parse shift time: {shift}")
        return None

    def _shifts_overlap(self, shift1: str, shift2: str) -> bool:
        """Check if two shifts overlap."""
        time1 = self._parse_shift_time(shift1)
        time2 = self._parse_shift_time(shift2)

        if not time1 or not time2:
            return False

        start1, end1 = time1
        start2, end2 = time2

        # Handle overnight shifts
        if end1 < start1:  # Shift 1 is overnight
            return not (end2 <= start1 and end2 <= start2)
        if end2 < start2:  # Shift 2 is overnight
            return not (end1 <= start2 and end1 <= start1)

        # Normal case
        return not (end1 <= start2 or end2 <= start1)

    def _generate_context(self) -> str:
        """Generate context for the LLM prompt."""
        if self.agents_df is None:
            return ""

        context = []

        # Basic statistics
        total_agents = len(self.agents_df['ID'].unique())
        total_dates = len(self.date_range)
        context.append(f"Total agents: {total_agents}")
        context.append(f"Date range: {self.date_range[0].date()} to {self.date_range[-1].date()} ({total_dates} days)")

        # Agents by queue
        queue_counts = self.agents_df.groupby('Queue')['ID'].nunique()
        context.append("\nAgents by queue:")
        for queue, count in queue_counts.items():
            context.append(f"- {queue}: {count} agents")

        # Agents by shift
        shift_counts = self.agents_df.groupby('Shift')['ID'].nunique()
        context.append("\nAgents by shift:")
        for shift, count in shift_counts.items():
            context.append(f"- {shift}: {count} agents")

        # Agents by batch
        batch_counts = self.agents_df.groupby('Batch')['ID'].nunique()
        context.append("\nAgents by batch (lower numbers = higher priority):")
        for batch, count in sorted(batch_counts.items()):
            context.append(f"- Batch {batch}: {count} agents")

        # Seat area capacities
        context.append("\nSeat area capacities:")
        for area, capacity in self.area_capacities.items():
            context.append(f"- {area}: {capacity} seats")

        return "\n".join(context)

    def _generate_prompt(self) -> str:
        """Generate the complete prompt for the LLM."""
        context = self._generate_context()

        prompt = f"""
You are an expert seat assignment system for a contact center. Your task is to assign seats to agents based on the following requirements and constraints.

**Context:**
{context}

**Seat Areas and Ranges:**
- OPS1-A: Seats 1-12
- OPS1-B: Seats 13-18
- OPS1-C: Seats 19-24
- OPS1-D: Seats 25-39
- OPS1-E: Seats 40-47
- OPS2: Seats 48-61 (Seat 61 is reserved)
- OPS3: Seats 62-69
- TRN: Seats 71-100 (training)

**Rules:**
{self.rules if self.rules else "No specific rules provided. Use general best practices for seat assignment."}

**General Assignment Guidelines:**
1. Priority-based grouping:
   - Group agents by shift first, then by supervisor, then by queue
   - Senior agents (lower batch numbers) should get priority
   - DH (Department Heads) have highest priority

2. Shift constraints:
   - Agents with non-overlapping shifts can share the same seat
   - When sharing, format as "morning_agent/afternoon_agent"

3. Proximity rules:
   - Agents from the same queue should be seated near each other when possible
   - Supervisors should be seated near their team members

4. Capacity constraints:
   - Do not exceed the capacity of any area
   - Try to distribute agents evenly across areas

5. Special cases:
   - Handle preferred/avoided seats if specified in rules
   - Respect any area constraints for specific queues/shifts

**Input Data Format:**
The input data contains these columns: ID, Name, Date, Queue, Shift, Supervisor, Batch

**Output Requirements:**
Return a JSON object with seat assignments with this structure:
{{
    "assignments": {{
        "YYYY-MM-DD": {{
            "seat_number": "agent_name",
            "seat_number": "agent1_name/agent2_name"  // for shared seats
        }}
    }},
    "unassigned": ["agent_name1", "agent_name2"],
    "notes": "Any important notes about the assignment"
}}

**Important Notes:**
- Only return the JSON object, no other text
- Make sure all seat numbers are valid (1-69, excluding reserved seats unless specified)
- If you can't assign a seat to an agent, add them to the unassigned list
- Be consistent with date formats (YYYY-MM-DD)
- For shared seats, always put the earlier shift first in the format "morning/afternoon"
"""
        return prompt

    def _call_ollama(self, prompt: str) -> Optional[Dict]:
        """Call Ollama using the official library."""
        try:
            response = self.client.generate(
                model=self.model,
                prompt=prompt,
                stream=False,
                options={
                    'temperature': 0.7,
                    'num_ctx': 4096
                }
            )

            if not response or 'response' not in response:
                logger.error("Invalid response from Ollama")
                return None

            # Parse the response
            try:
                json_start = response['response'].find('{')
                json_end = response['response'].rfind('}') + 1
                if json_start == -1 or json_end == 0:
                    logger.error("No valid JSON found in response")
                    return None

                json_str = response['response'][json_start:json_end]
                return json.loads(json_str)
            except json.JSONDecodeError as e:
                logger.error(f"Error parsing JSON response: {e}")
                return None
        except Exception as e:
            logger.error(f"Error calling Ollama: {e}")
            return None

    def _fallback_assignment(self) -> Dict:
        """Fallback assignment algorithm when LLM fails."""
        logger.warning("Using fallback assignment algorithm")

        assignments = {date.strftime('%Y-%m-%d'): {} for date in self.date_range}
        unassigned = []

        # Sort agents by priority (batch, then shift)
        sorted_agents = self.agents_df.sort_values(['Batch', 'Shift', 'Queue'])

        # Assign seats sequentially
        seat_index = 1
        for date in self.date_range:
            date_str = date.strftime('%Y-%m-%d')
            daily_agents = sorted_agents[sorted_agents['Date'] == date]

            for _, agent in daily_agents.iterrows():
                # Skip reserved seats
                while seat_index in RESERVED_SEATS:
                    seat_index += 1
                    if seat_index > TOTAL_SEATS:
                        seat_index = 1

                # Check if seat is already assigned for this date
                if seat_index in assignments[date_str]:
                    # Try to find another seat
                    for s in range(1, TOTAL_SEATS + 1):
                        if s not in RESERVED_SEATS and s not in assignments[date_str]:
                            seat_index = s
                            break
                    else:
                        unassigned.append(agent['Name'])
                        continue

                assignments[date_str][seat_index] = agent['Name']
                seat_index += 1
                if seat_index > TOTAL_SEATS:
                    seat_index = 1

        return {
            "assignments": assignments,
            "unassigned": unassigned,
            "notes": "Fallback assignment algorithm used. Assignments may not follow all rules."
        }

    def assign_seats(self) -> Optional[Dict]:
        """Main method to assign seats using Ollama."""
        if not self.validate_connection():
            logger.error("Cannot proceed without working Ollama connection")
            return None

        prompt = self._generate_prompt()
        logger.debug(f"Generated prompt (first 500 chars): {prompt[:500]}...")

        # Call Ollama
        response = self._call_ollama(prompt)
        if not response:
            logger.warning("LLM call failed, using fallback assignment")
            return self._fallback_assignment()

        # Validate response structure
        if "assignments" not in response or not isinstance(response["assignments"], dict):
            logger.warning("Invalid response structure, using fallback assignment")
            return self._fallback_assignment()

        return response

    def _optimize_shared_seats(self, assignments: Dict) -> Dict:
        """Optimize assignments to maximize seat sharing for non-overlapping shifts."""
        optimized = {date: {} for date in assignments.get("assignments", {})}
        unassigned = assignments.get("unassigned", [])

        for date_str, date_assignments in assignments.get("assignments", {}).items():
            try:
                date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                logger.warning(f"Invalid date format: {date_str}")
                continue

            daily_agents = self.agents_df[self.agents_df['Date'] == pd.Timestamp(date)]
            seat_map = {}

            # First pass: collect all assignments
            for seat, agent_name in date_assignments.items():
                try:
                    seat_num = int(seat)
                    if seat_num not in seat_map:
                        seat_map[seat_num] = []
                    seat_map[seat_num].append(agent_name)
                except ValueError:
                    continue

            # Second pass: optimize sharing
            for seat, agents in seat_map.items():
                if len(agents) == 1:
                    optimized[date_str][seat] = agents[0]
                    continue

                # Find all agents assigned to this seat
                seat_agents = daily_agents[daily_agents['Name'].isin(agents)]

                # Group by shift and check for overlaps
                shift_groups = {}
                for _, agent in seat_agents.iterrows():
                    shift = agent['Shift']
                    if shift not in shift_groups:
                        shift_groups[shift] = []
                    shift_groups[shift].append(agent['Name'])

                # Check if all shifts are non-overlapping
                shifts = list(shift_groups.keys())
                can_share = True

                for i in range(len(shifts)):
                    for j in range(i + 1, len(shifts)):
                        if self._shifts_overlap(shifts[i], shifts[j]):
                            can_share = False
                            break
                    if not can_share:
                        break

                if can_share:
                    # Sort by shift start time
                    sorted_shifts = sorted(
                        shifts,
                        key=lambda s: self._parse_shift_time(s)[0] if self._parse_shift_time(s) else time(0, 0)
                    )
                    optimized[date_str][seat] = "/".join(
                        shift_groups[s][0] for s in sorted_shifts
                    )
                else:
                    # Can't share, assign to first agent and unassign others
                    optimized[date_str][seat] = shift_groups[shifts[0]][0]
                    for shift in shifts[1:]:
                        unassigned.extend(shift_groups[shift])

        return {
            "assignments": optimized,
            "unassigned": unassigned,
            "notes": assignments.get("notes", "") + "\nOptimized seat sharing for non-overlapping shifts"
        }

    def generate_output(self, assignments: Dict, output_path: str) -> bool:
        """Generate the output Excel file with seat assignments."""
        try:
            wb = Workbook()

            # Sheet 1: Seat Assignments
            ws_assignments = wb.active
            ws_assignments.title = "Seat Assignments"

            # Prepare data
            all_dates = sorted(assignments['assignments'].keys())
            all_seats = range(1, TOTAL_SEATS + 1)

            # Write headers
            ws_assignments.append(['Area', 'Seat'] + all_dates)

            # Write seat assignments
            for area, seats in SEAT_AREAS.items():
                for seat in seats:
                    if seat in RESERVED_SEATS:
                        continue

                    row = [area, seat]
                    for date in all_dates:
                        assignment = assignments['assignments'][date].get(seat, "")
                        row.append(assignment)
                    ws_assignments.append(row)

            # Format the assignments sheet
            self._format_assignments_sheet(ws_assignments, all_dates)

            # Sheet 2: Summary
            ws_summary = wb.create_sheet("Summary")
            self._create_summary_sheet(ws_summary, assignments, all_dates)

            # Sheet 3: Legend
            ws_legend = wb.create_sheet("Legend")
            self._create_legend_sheet(ws_legend)

            # Save the workbook
            wb.save(output_path)
            logger.info(f"Output saved to {output_path}")
            return True
        except Exception as e:
            logger.error(f"Error generating output: {e}")
            return False

    def _format_assignments_sheet(self, ws, dates: List[str]):
        """Format the assignments sheet."""
        ws.freeze_panes = 'C2'

        # Set column widths
        ws.column_dimensions['A'].width = 10  # Area
        ws.column_dimensions['B'].width = 6   # Seat
        for i, date in enumerate(dates, start=3):
            ws.column_dimensions[get_column_letter(i)].width = 20

        # Header formatting
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Center align all cells
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center')

    def _create_summary_sheet(self, ws, assignments: Dict, dates: List[str]):
        """Create the summary sheet."""
        total_agents = len(self.agents_df['ID'].unique())
        total_dates = len(dates)
        total_seats = TOTAL_SEATS - len(RESERVED_SEATS)
        unassigned_count = len(assignments['unassigned'])

        # Calculate seat utilization
        seat_usage = {date: len(assignments['assignments'][date]) for date in dates}
        avg_seat_usage = sum(seat_usage.values()) / len(seat_usage)
        max_seat_usage = max(seat_usage.values())
        min_seat_usage = min(seat_usage.values())

        # Calculate shared seats
        shared_seats = 0
        for date in dates:
            shared_seats += sum(
                1 for assignment in assignments['assignments'][date].values()
                if '/' in str(assignment)
            )

        # Write summary data
        ws.append(["Seat Assignment Summary"])
        ws.append([])
        ws.append(["General Information"])
        ws.append(["Total agents", total_agents])
        ws.append(["Total dates", total_dates])
        ws.append(["Date range", f"{dates[0]} to {dates[-1]}"])
        ws.append(["Total available seats", total_seats])
        ws.append(["Reserved seats", len(RESERVED_SEATS)])
        ws.append(["Unassigned agents", unassigned_count])
        ws.append(["Percentage assigned", f"{(1 - unassigned_count/total_agents)*100:.1f}%"])
        ws.append([])
        ws.append(["Seat Utilization"])
        ws.append(["Average seats used per day", f"{avg_seat_usage:.1f}"])
        ws.append(["Maximum seats used in a day", max_seat_usage])
        ws.append(["Minimum seats used in a day", min_seat_usage])
        ws.append(["Total shared seats", shared_seats])
        ws.append(["Average shared seats per day", f"{shared_seats/total_dates:.1f}"])
        ws.append([])
        ws.append(["Notes"])
        ws.append([assignments['notes']])

        # Format the sheet
        for cell in ws[1]:
            cell.font = Font(bold=True, size=14)

        for cell in ws[3]:
            cell.font = Font(bold=True)

        for cell in ws[12]:
            cell.font = Font(bold=True)

        for cell in ws[17]:
            cell.font = Font(bold=True)

    def _create_legend_sheet(self, ws):
        """Create the legend sheet."""
        ws.append(["Seat Assignment Legend"])
        ws.append([])
        ws.append(["Format Explanation"])
        ws.append(["- Each row represents a seat in a specific area"])
        ws.append(["- Each column represents a date"])
        ws.append(["- Cell content shows the agent(s) assigned to that seat on that date"])
        ws.append(["- For shared seats (non-overlapping shifts), format is 'agent1/agent2'"])
        ws.append(["- Empty cells mean the seat is unassigned for that date"])
        ws.append([])
        ws.append(["Seat Areas"])
        for area, seats in SEAT_AREAS.items():
            ws.append([f"- {area}: Seats {min(seats)}-{max(seats)}"])
        ws.append(["- Reserved seat: 61"])
        ws.append([])
        ws.append(["Shift Times"])
        for shift, times in SHIFT_TIMES.items():
            ws.append([f"- {shift}: {times[0].strftime('%H:%M')}-{times[1].strftime('%H:%M')}"])

        # Format the sheet
        for cell in ws[1]:
            cell.font = Font(bold=True, size=14)

        for cell in ws[3]:
            cell.font = Font(bold=True)

        for cell in ws[10]:
            cell.font = Font(bold=True)

        for cell in ws[18]:
            cell.font = Font(bold=True)

def main():
    parser = argparse.ArgumentParser(
        description="AI-Powered Seat Assignment System for Contact Centers"
    )
    parser.add_argument(
        "--agents",
        required=True,
        help="Path to Excel file containing agent schedules"
    )
    parser.add_argument(
        "--rules",
        help="Path to text file containing assignment rules (optional)"
    )
    parser.add_argument(
        "--output",
        help="Output file path (optional, defaults to timestamped filename)"
    )
    parser.add_argument(
        "--ollama-host",
        required=True,
        help="Host address of the Ollama server (e.g., http://localhost:11434)"
    )
    parser.add_argument(
        "--model",
        default="llama3",
        help="Ollama model to use (default: llama3)"
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Enable verbose logging"
    )

    args = parser.parse_args()

    if args.verbose:
        logger.setLevel(logging.DEBUG)

    # Initialize the seat assignment system
    system = SeatAssignmentSystem(
        ollama_host=args.ollama_host,
        model=args.model
    )

    # Load agent data
    if not system.load_agent_data(args.agents):
        logger.error("Failed to load agent data. Exiting.")
        sys.exit(1)

    # Load rules
    if not system.load_rules(args.rules):
        logger.error("Failed to load rules. Exiting.")
        sys.exit(1)

    # Assign seats
    assignments = system.assign_seats()
    if not assignments:
        logger.error("Failed to generate seat assignments. Exiting.")
        sys.exit(1)

    # Optimize shared seats
    optimized = system._optimize_shared_seats(assignments)

    # Generate output filename if not provided
    if not args.output:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        input_stem = Path(args.agents).stem
        output_path = f"{input_stem}_{timestamp}_seating_arrangement.xlsx"
    else:
        output_path = args.output

    # Generate output
    if not system.generate_output(optimized, output_path):
        logger.error("Failed to generate output file. Exiting.")
        sys.exit(1)

if __name__ == "__main__":
    import re  # Add this import at the top
    main()