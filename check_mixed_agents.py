import pandas as pd
import openpyxl

df = pd.read_excel('Agent Schedules (Dec 1st - 21st)_20251204_071622_clean.xlsx')
mixed = df[df['Shift'].isin(['10AM', '11AM', '2PM'])]
print(f'Total mixed shift agents: {len(mixed)}')
print(f'Unique mixed shift dates: {mixed["Date"].nunique()}')
print(f'Unique mixed shift agents: {mixed["Name"].nunique()}')
print()
print('Sample mixed shift agents:')
for idx, row in mixed.head(10).iterrows():
    print(f'  {row["Date"]}: {row["Name"]} ({row["Shift"]})')

# Check Mixed Seating sheet for data
print("\n" + "=" * 70)
print("Mixed Seating sheet sample (looking for mixed shift agents):")
wb = openpyxl.load_workbook('Agent Schedules (Dec 1st - 21st)_20251218_151404_seating_arrangement.xlsx')
ws = wb['Mixed Seating']
found_count = 0
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3):
    for cell in row:
        if cell.value and cell.value != '':
            print(f"  Found: {cell.value}")
            found_count += 1
            if found_count >= 10:
                break
    if found_count >= 10:
        break

if found_count == 0:
    print("  (No mixed shift agents in this date range)")
